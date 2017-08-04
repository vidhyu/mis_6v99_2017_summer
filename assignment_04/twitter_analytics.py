#Assignment02 

from __future__ import print_function
import requests # for sending Http requests
import os       # for getting functions to set OS paths
import zipfile  # module for Zip and Unzip file functions
import openpyxl # create/delete xl files. openpyxl is a pure python reader and writer of Excel OpenXML files (`.xlsx`, `.xlsm`).
import sqlite3  # for sqlite DB
import glob     # glob patterns specify sets of filenames with wildcard characters
import getpass
import csv      # To work with CSV files
  
# Step1 : Downloading and Extracting CSV files from data.medicare.gov

url = "https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip"

r = requests.get(url) # this will connect to the url

staging_dir_name = "staging"

os.mkdir(staging_dir_name)  # mkdir will make the directory if it doesn't exist in the path

zip_file_name = os.path.join(staging_dir_name, "Hospital_Data.zip") # Zip file will be created in the mentioned Staging path

zf = open(zip_file_name, "wb") # The wb indicates that the file is opened for writing in binary mode. On Unix systems (Linux, Mac OS X, etc.), binary mode does nothing - they treat text files the same way that any other files are treated. On Windows, however, text files are written with slightly modified line endings. This causes a serious problem when dealing with actual binary files, like exe or jpg files. Therefore, when opening files which are not supposed to be text, even in Unix, you should use wb or rb. Use plain w or r only for text files.

zf.write(r.content) # writes the entire content in the url to the zip file

zf.close() # closes the file

z = zipfile.ZipFile(zip_file_name,'r') # Open a ZIP file, where file can be either a path to a file (a string) or a file-like object. The mode parameter should be 'r' to read an existing file, 'w' to truncate and write a new file, or 'a' to append to an existing file.

z.extractall(staging_dir_name) 

z.close()


# Step2 : Create a SQLite DB to fill the data

glob_dir = os.path.join(staging_dir_name,"*.csv")  # to get list of csv files
valid_characters = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'


conn = sqlite3.connect("medicare_hospital_compare.db") #connection to open db, if the db doesn't exist it creates a new one.
conn.text_factory = str
c1 = conn.cursor()  # Once you have a Connection, you can create a Cursor object and call its execute() method to perform SQL commands


for file_name in glob.glob(glob_dir):
    tablename = os.path.splitext(os.path.basename(file_name))[0]
    if not(tablename == "FY2015_Percent_Change_in_Medicare_Payments"):
    # remove the path and extension and use what's left as a table name
        
        tablename = tablename.replace(' ', '_').replace('-','_').replace('%','pct').replace('/','_').lower().replace('''""''', '_')
        if tablename[0] not in valid_characters:
            tablename = "t_"+ tablename
    
        with open(file_name, "rt", encoding = 'cp1252') as f:
            reader = csv.reader(f, delimiter=',', skipinitialspace=True)
            header = next(reader)
            header = ','.join(header)
            header = header.replace('"', '').replace(' ', '_').replace('-','_').replace('%','pct').replace('/','_').lower().replace('''""''', '_')   
            header = header.split(',')
            header = [x if x[0] in valid_characters else 'c_'+ x for x in header]

            sql = "DROP TABLE IF EXISTS '" + tablename + "'"
            c1.execute(sql)
            sql = "CREATE TABLE " + tablename + "(" + ",".join(column+ " text" for column in header) + ")"
            c1.execute(sql)
    
    
    
    
            for row in reader:
                insertsql = "INSERT INTO %s VALUES (%s)" % (tablename,
                                               ", ".join([ "?" for column in row ]))
                if len(row) != 1:
                    c1.execute(insertsql, row)
                    
        conn.commit()

c1.close()
conn.close()
    
    
    
 #step 4 : Creating measures_statistics.xlsx


#creating a workbook
wb3 = openpyxl.Workbook()


#creating first sheet and inserting values
sheet_1 = wb3.create_sheet("Nationwide") 

sql = "select measure_id, measure_name, min(score) as minimum, max(score) as maximum, avg(score) as Average from timely_and_effective_care___hospital where length(score) < 5 group by measure_id, measure_name;"


result = [c.execute(sql)]
db_data = []
for row in result:
    for column in row:
        data = [column]
        db_data = db_data + data

        
i = len(db_data)

#creating first column and inserting values
sheet_1.cell(row = 1, column = 1, value = "Measure ID")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=1, value = db_data[x-2][0])

#creating second column and inserting values
sheet_1.cell(row = 1, column = 2, value = "Measure Name")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=2, value = db_data[x-2][1])
    

#creating third column and inserting values
sheet_1.cell(row = 1, column = 3, value = "Minimum")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=3, value = db_data[x-2][2])
    

#creating fourth column and inserting values
sheet_1.cell(row = 1, column = 4, value = "Maximum")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=4, value = db_data[x-2][3])
    
    
#creating fifth column and inserting values
sheet_1.cell(row = 1, column = 5, value = "Average")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=5, value = db_data[x-2][4])

  
#fetching data for calculating std dev

sql = "select measure_id, score from timely_and_effective_care___hospital where length(score) < 5 order by measure_id;"

result = [c.execute(sql)]
std_data = []
for row in result:
    for column in row:
        data = list(column)
        std_data = std_data + [data]
        
        
for i in range(len(std_data)):
    std_data[i][1] = int(float(std_data[i][1]))
    

std_dict = {k:list(x[1] for x in v) for k,v in it.groupby(sorted(std_data), key=lambda x: x[0])}

std_dev = []
for j in std_dict:
    arr = np.array(std_dict[j])
    x = np.std(arr, axis=0)
    std_dev = std_dev + [x]

i = len(std_dev)    
#creating sixth column and inserting values
sheet_1.cell(row = 1, column = 6, value = "Standard Deviation")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=6, value = std_dev[x-2])
    


    
#creating sheets with state names and inserting values
for x in state_names:
    z = wb3.create_sheet(state_names[x])
    
    sql = "select measure_id, measure_name, min(score) as minimum, max(score) as maximum, avg(score) as Average from timely_and_effective_care___hospital where state = '" + x + "' and length(score) < 5 group by measure_id, measure_name;"
         
    result = [c.execute(sql)]
    db_data = []
    for row in result:
        for column in row:
            data = [column]
            db_data = db_data + data
    
    
    i = len(db_data)

    #creating first column and inserting values
    z.cell(row = 1, column = 1, value = "Measure ID")

    for y in range(2,i+2):
        z.cell(row = y, column=1, value = db_data[y-2][0])

#creating second column and inserting values
    z.cell(row = 1, column = 2, value = "Measure Name")

    for y in range(2,i+2):
        z.cell(row = y, column=2, value = db_data[y-2][1])
    

#creating third column and inserting values
    z.cell(row = 1, column = 3, value = "Minimum")

    for y in range(2,i+2):
        z.cell(row = y, column=3, value = db_data[y-2][2])
    

#creating fourth column and inserting values
    z.cell(row = 1, column = 4, value = "Maximum")

    for y in range(2,i+2):
        z.cell(row = y, column=4, value = db_data[y-2][3])
    
    
#creating fifth column and inserting values
    z.cell(row = 1, column = 5, value = "Average")

    for y in range(2,i+2):
        z.cell(row = y, column=5, value = db_data[y-2][4])

    
#creating sixth column and inserting values
    sql = "select measure_id, score from timely_and_effective_care___hospital where state = '" + x + "' and length(score) < 5 order by measure_id;"

    result = [c.execute(sql)]
    std_data = []
    for row in result:
        for column in row:
            data = list(column)
            std_data = std_data + [data]
        
        
    for i in range(len(std_data)):
        std_data[i][1] = int(float(std_data[i][1]))
    

    std_dict = {k:list(x[1] for x in v) for k,v in it.groupby(sorted(std_data), key=lambda x: x[0])}

    std_dev = []
    for j in std_dict:
        arr = np.array(std_dict[j])
        x = np.std(arr, axis=0)
        std_dev = std_dev + [x]
        
    i = len(std_dev)
    #creating sixth column and inserting values
    z.cell(row = 1, column = 6, value = "Standard Deviation")

    for x in range(2,i+2):
        z.cell(row = x, column=6, value = std_dev[x-2])
    
    
      
    
#remove sheet in xl
wb3.remove_sheet(wb3.get_sheet_by_name('Sheet'))

wb3.save("measures_statistics.xlsx")

wb3.close()


conn.commit()

c.close()
conn.close()   
    
    
    
    
    
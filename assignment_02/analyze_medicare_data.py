#Assignment02 

from __future__ import print_function
import requests # for sending Http requests
import os       # for getting functions to set OS paths
import zipfile  # module for Zip and Unzip file functions
import openpyxl # create/delete xl files. openpyxl is a pure python reader and writer of Excel OpenXML files (`.xlsx`, `.xlsm`).
import sqlite3  # for sqlite DB
import glob     # glob patterns specify sets of filenames with wildcard characters
import itertools as it  
import csv      # To work with CSV files
import numpy as np
 
# Step1 : Downloading and Extracting CSV files from data.medicare.gov

url = "https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip"

r = requests.get(url) # this will connect to the url

staging_dir_name = "staging" #directory name

os.mkdir(staging_dir_name)  # mkdir will make the directory if it doesn't exist in the path

zip_file_name = os.path.join(staging_dir_name, "Hospital_Data.zip") # Zip file will be created in the mentioned Staging path

zf = open(zip_file_name, "wb") # The wb indicates that the file is opened for writing in binary mode. On Unix systems (Linux, Mac OS X, etc.), binary mode does nothing - they treat text files the same way that any other files are treated. On Windows, however, text files are written with slightly modified line endings. This causes a serious problem when dealing with actual binary files, like exe or jpg files. Therefore, when opening files which are not supposed to be text, even in Unix, you should use wb or rb. Use plain w or r only for text files.

zf.write(r.content) # writes the entire content in the url to the zip file

zf.close() # closes the file

z = zipfile.ZipFile(zip_file_name,'r') # Open a ZIP file, where file can be either a path to a file (a string) or a file-like object. The mode parameter should be 'r' to read an existing file, 'w' to truncate and write a new file, or 'a' to append to an existing file.

z.extractall(staging_dir_name) #extracts all the files in staging

z.close()

# Step2 : Create a SQLite DB to fill the data

glob_dir = os.path.join(staging_dir_name,"*.csv")  # to get list of csv files
valid_characters = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
conn = sqlite3.connect("medicare_hospital_compare.db") #connection to open db, if the db doesn't exist it creates a new one.
conn.text_factory = str
c1 = conn.cursor()  # Once you have a Connection, you can create a Cursor object and call its execute() method to perform SQL commands

for file_name in glob.glob(glob_dir):
    tablename = os.path.splitext(os.path.basename(file_name))[0] # remove the path and extension and use what's left as a table name
    if not(tablename == "FY2015_Percent_Change_in_Medicare_Payments"):  #removing the corrupt file from iterations
    
        tablename = tablename.replace(' ', '_').replace('-','_').replace('%','pct').replace('/','_').lower().replace('''""''', '_')
        if tablename[0] not in valid_characters:   #cross-checking the first letter of tablename to be in valid_characters
            tablename = "t_"+ tablename
    
        with open(file_name, "rt", encoding = 'cp1252') as f:
            reader = csv.reader(f, delimiter=',', skipinitialspace=True)
            header = next(reader)  #takes the first row as header
            header = ','.join(header)
            header = header.replace('"', '').replace(' ', '_').replace('-','_').replace('%','pct').replace('/','_').lower().replace('''""''', '_')   
            header = header.split(',')
            header = [x if x[0] in valid_characters else 'c_'+ x for x in header]
            #checking for existing tablenames and create tables
            sql = "DROP TABLE IF EXISTS '" + tablename + "'"
            c1.execute(sql)
            sql = "CREATE TABLE " + tablename + "(" + ",".join(column+ " text" for column in header) + ")"
            c1.execute(sql)
            #inserting values in created tables using the other data after stripping the header
            for row in reader:
                insertsql = "INSERT INTO %s VALUES (%s)" % (tablename,
                                               ", ".join([ "?" for column in row ]))
                if len(row) != 1:           #if the end of the file has blank rows, the insert query fails because blank values cannot be put into all columns
                    c1.execute(insertsql, row)
                    
        conn.commit()

c1.close()
conn.close()
    
#step 3 : hospital ranking excel


k_url = "http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"

r = requests.get(k_url)

xf = open("hospital_ranking_focus_states.xlsx", "wb")

xf.write(r.content)

xf.close()

wb = openpyxl.load_workbook("hospital_ranking_focus_states.xlsx")
    
sheet1 = wb.get_sheet_by_name("Hospital National Ranking")


i = 1
provider_id = {}
while sheet1.cell(row = i+1, column = 1).value !=None:
    line = {sheet1.cell(row = i+1, column = 1).value: sheet1.cell(row = i+1, column = 2).value}
    provider_id = {**provider_id, **line}
    i +=1
    
empty_list = []
for x in provider_id:
    test = ["'" + x + "'"]
    empty_list = empty_list + test
final_list = empty_list[:100]
final_list = ','.join(final_list)

sheet2 = wb.get_sheet_by_name("Focus States")
state_names = {}
i = 1
while sheet2.cell(row = i+1, column = 1).value !=None:
    line = {sheet2.cell(row = i+1, column = 2).value : sheet2.cell(row = i+1, column = 1).value}
    state_names = {**state_names, **line} #each * indicate offset and value
    i +=1
    
#creating a workbook
wb2 = openpyxl.Workbook()


#creating first sheet and inserting values
sheet_1 = wb2.create_sheet("Nationwide")

conn = sqlite3.connect("medicare_hospital_compare.db")
c = conn.cursor() 

sql = "select provider_id, hospital_name, city, state, county_name from hospital_general_information where provider_id in (" + final_list + ")"             
result = [c.execute(sql)]
db_data = []
for row in result:
    for column in row:
        data = [column]
        db_data = db_data + data

        
rank = []
for row in db_data:
    dummy = list(row) + [provider_id[row[0]]]
    rank = rank + [dummy]
    
sort_rank = sorted(rank, key=lambda k: k[5])

rank_removed = []
for row in sort_rank:
    y = row[:5]
    rank_removed = rank_removed + [y]
    
    
i = len(rank_removed)

#creating first column and inserting values
sheet_1.cell(row = 1, column = 1, value = "Provider ID")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=1, value = rank_removed[x-2][0])

#creating second column and inserting values
sheet_1.cell(row = 1, column = 2, value = "Hospital Name")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=2, value = rank_removed[x-2][1])
    

#creating third column and inserting values
sheet_1.cell(row = 1, column = 3, value = "City")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=3, value = rank_removed[x-2][2])
    

#creating fourth column and inserting values
sheet_1.cell(row = 1, column = 4, value = "State")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=4, value = rank_removed[x-2][3])
    
    
#creating fifth column and inserting values
sheet_1.cell(row = 1, column = 5, value = "County")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=5, value = rank_removed[x-2][4])
    

#creating sheets with state names and inserting values
for x in state_names:
    z = wb2.create_sheet(state_names[x])
    
    sql = "select provider_id, hospital_name, city, state, county_name from hospital_general_information where state = '" + x + "'"          
    result = [c.execute(sql)]
    db_data = []
    for row in result:
        for column in row:
            data = [column]
            db_data = db_data + data

    
    rank = []
    for row in db_data:
        dummy = list(row) + [provider_id[row[0]]]
        rank = rank + [dummy]
    
    sort_rank = sorted(rank, key=lambda k: k[5])

    rank_removed = []
    for row in sort_rank:
        y = row[:5]
        rank_removed = rank_removed + [y]
    
    
    i = len(rank_removed[:100])

    #creating first column and inserting values
    z.cell(row = 1, column = 1, value = "Provider ID")

    for y in range(2,i+2):
        z.cell(row = y, column=1, value = rank_removed[y-2][0])

    #creating second column and inserting values
    z.cell(row = 1, column = 2, value = "Hospital Name")

    for y in range(2,i+2):
        z.cell(row = y, column=2, value = rank_removed[y-2][1])
    

    #creating third column and inserting values
    z.cell(row = 1, column = 3, value = "City")

    for y in range(2,i+2):
        z.cell(row = y, column=3, value = rank_removed[y-2][2])
    

    #creating fourth column and inserting values
    z.cell(row = 1, column = 4, value = "State")

    for y in range(2,i+2):
        z.cell(row = y, column=4, value = rank_removed[y-2][3])
    
    
    #creating fifth column and inserting values
    z.cell(row = 1, column = 5, value = "County")

    for y in range(2,i+2):
        z.cell(row = y, column=5, value = rank_removed[y-2][4])
    

    
#remove sheet in xl
wb2.remove_sheet(wb2.get_sheet_by_name('Sheet'))

wb2.save("hospital_ranking.xlsx")

wb2.close()



#step 4 : Creating measures_statistics.xlsx


#creating a workbook
wb3 = openpyxl.Workbook()


#creating first sheet and inserting values
sheet_1 = wb3.create_sheet("Nationwide") 

sql = "select measure_id, measure_name, score from timely_and_effective_care___hospital where length(score) < 5 order by measure_id;"


result = [c.execute(sql)]
db_data = []
for row in result:
    for column in row:
        data = list(column)
        db_data = db_data + [data]

for i in range(len(db_data)):
    db_data[i][2] = int(float(db_data[i][2]))
    

data_dict = {k:list(x[2] for x in v) for k,v in it.groupby(sorted(db_data), key=lambda x: x[0])}

measure_dict = {k:list(x[1] for x in v) for k,v in it.groupby(sorted(db_data), key=lambda x: x[0])}

#to get measure_ids in list

measure_ids = list(measure_dict.keys())

#to get names in list
measure_name = []
for x in measure_ids:
    t = list(measure_dict[x])
    measure_name = measure_name + [t[0]]

#to get min in list
min_score = []
for j in data_dict:
    arr = np.array(data_dict[j])
    x = np.min(arr, axis=0)
    min_score = min_score + [x]

#to get max in list    
max_score = []
for j in data_dict:
    arr = np.array(data_dict[j])
    x = np.max(arr, axis=0)
    max_score = max_score + [x]

#to get avg in list   
avg_score = []
for j in data_dict:
    arr = np.array(data_dict[j])
    x = np.average(arr, axis=0)
    avg_score = avg_score + [x]
    
i = len(data_dict)

#creating first column and inserting values
sheet_1.cell(row = 1, column = 1, value = "Measure ID")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=1, value = measure_ids[x-2])

#creating second column and inserting values
sheet_1.cell(row = 1, column = 2, value = "Measure Name")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=2, value = measure_name[x-2])
    

#creating third column and inserting values
sheet_1.cell(row = 1, column = 3, value = "Minimum")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=3, value = min_score[x-2])
    

#creating fourth column and inserting values
sheet_1.cell(row = 1, column = 4, value = "Maximum")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=4, value = max_score[x-2])
    
    
#creating fifth column and inserting values
sheet_1.cell(row = 1, column = 5, value = "Average")

for x in range(2,i+2):
    sheet_1.cell(row = x, column=5, value = avg_score[x-2])

  
#fetching data for calculating std dev

sql = "select measure_id, score from timely_and_effective_care___hospital where length(score) < 5 order by measure_id;"

result = [c.execute(sql)]
std_data = []
for row in result:
    for column in row:
        data = list(column)
        std_data = std_data + [data]
        
        
for i in range(len(std_data)):
    std_data[i][1] = int(float(std_data[i][1]))
    

std_dict = {k:list(x[1] for x in v) for k,v in it.groupby(sorted(std_data), key=lambda x: x[0])}

std_dev = []
for j in std_dict:
    arr = np.array(std_dict[j])
    x = np.std(arr, axis=0)
    std_dev = std_dev + [x]

i = len(std_dev)    
#creating sixth column and inserting values
sheet_1.cell(row = 1, column = 6, value = "Standard Deviation")


for x in range(2,i+2):
    sheet_1.cell(row = x, column=6, value = std_dev[x-2])
    


    
#creating sheets with state names and inserting values
for l in state_names:
    sheet_1 = wb3.create_sheet(state_names[l])
    
    sql = "select measure_id, measure_name, score from timely_and_effective_care___hospital where state = '" + l + "' and length(score) < 5 order by measure_id;"
         
    result = [c.execute(sql)]
    db_data = []
    for row in result:
        for column in row:
            data = list(column)
            db_data = db_data + [data]

    for i in range(len(db_data)):
        db_data[i][2] = int(float(db_data[i][2]))
    

    data_dict = {k:list(x[2] for x in v) for k,v in it.groupby(sorted(db_data), key=lambda x: x[0])}

    measure_dict = {k:list(x[1] for x in v) for k,v in it.groupby(sorted(db_data), key=lambda x: x[0])}

    #to get measure_ids in list

    measure_ids = list(measure_dict.keys())
    
    #to get names in list

    measure_name = []
    for x in measure_ids:
        t = list(measure_dict[x])
        measure_name = measure_name + [t[0]]
        
    #to get min_scores in list
    min_score = []
    for j in data_dict:
        arr = np.array(data_dict[j])
        x = np.min(arr, axis=0)
        min_score = min_score + [x]
        
    #to get max in list
    max_score = []
    for j in data_dict:
        arr = np.array(data_dict[j])
        x = np.max(arr, axis=0)
        max_score = max_score + [x]
    
    #to get avg in list
    avg_score = []
    for j in data_dict:
        arr = np.array(data_dict[j])
        x = np.average(arr, axis=0)
        avg_score = avg_score + [x]
    
    
    
    i = len(data_dict)

    #creating first column and inserting values
    sheet_1.cell(row = 1, column = 1, value = "Measure ID")

    for x in range(2,i+2):
        sheet_1.cell(row = x, column=1, value = measure_ids[x-2])

    #creating second column and inserting values
    sheet_1.cell(row = 1, column = 2, value = "Measure Name")

    for x in range(2,i+2):
        sheet_1.cell(row = x, column=2, value = measure_name[x-2])
    

    #creating third column and inserting values
    sheet_1.cell(row = 1, column = 3, value = "Minimum")

    for x in range(2,i+2):
        sheet_1.cell(row = x, column=3, value = min_score[x-2])
    

    #creating fourth column and inserting values
    sheet_1.cell(row = 1, column = 4, value = "Maximum")

    for x in range(2,i+2):
        sheet_1.cell(row = x, column=4, value = max_score[x-2])
    
    
    #creating fifth column and inserting values
    sheet_1.cell(row = 1, column = 5, value = "Average")

    for x in range(2,i+2):
        sheet_1.cell(row = x, column=5, value = avg_score[x-2])

    
#creating sixth column and inserting values
    sql = "select measure_id, score from timely_and_effective_care___hospital where state = '" + l + "' and length(score) < 5 order by measure_id;"
     #storing data in list
    result = [c.execute(sql)]
    std_data = []
    for row in result:
        for column in row:
            data = list(column)
            std_data = std_data + [data]
        
        
    for i in range(len(std_data)):
        std_data[i][1] = int(float(std_data[i][1]))
    
    #creating dictionary with values
    std_dict = {k:list(x[1] for x in v) for k,v in it.groupby(sorted(std_data), key=lambda x: x[0])}

    std_dev = []
    for j in std_dict:
        arr = np.array(std_dict[j])
        x = np.std(arr, axis=0)
        std_dev = std_dev + [x]
    
    i = len(std_dev)
    #creating sixth column and inserting values
    sheet_1.cell(row = 1, column = 6, value = "Standard Deviation")

    for x in range(2,i+2):
        sheet_1.cell(row = x, column=6, value = std_dev[x-2])
    
    
    
    
    
    
#remove sheet in xl
wb3.remove_sheet(wb3.get_sheet_by_name('Sheet'))

wb3.save("measures_statistics.xlsx")

wb3.close()

conn.commit()

c.close()
conn.close() 